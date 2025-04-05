import base64
import time
from urllib.error import HTTPError

from selenium import webdriver
from selenium.common import UnexpectedAlertPresentException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import os
import openpyxl
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
import pygetwindow


def fetch_amazon_data(url):
    result = False
    title = ""
    keys = []
    try:

        options = Options()
        """
        cmd窗口执行以下命令,需要提前创建文件夹
        chrome.exe --remote-debugging-port=9527 --user-data-dir="F:\selenium\AutomationProfile"
        """
        options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
        driver = webdriver.Chrome(options=options)
        driver.execute_cdp_cmd(
            'Page.addScriptToEvaluateOnNewDocument',
            {'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'}
        )
        driver.get(url)


        """  
        location = driver.find_element(By.CSS_SELECTOR,"#glow-ingress-block")
        location.click()
        GLUXZipUpdateInput = driver.find_element(By.CSS_SELECTOR,"#GLUXZipUpdateInput")
        GLUXZipUpdateInput.send_keys("90091")
        
    
        change = driver.find_element(By.CSS_SELECTOR,"#GLUXChangePostalCodeLink")
        change.click()
        value = driver.find_element(By.CSS_SELECTOR,"#GLUXZipConfirmationValue")
        value.send_keys("90091")
        apply = driver.find_element(By.CSS_SELECTOR, "#GLUXZipUpdate-announce")
        apply.click()
        """
        title = driver.find_element(By.ID, "productTitle")
        keys = driver.find_elements(By.CSS_SELECTOR, "#feature-bullets span")
        result = True
        #driver.close()
    except HTTPError:
        print("HTTPError")
    except Exception:
        print("请检查url:"+url)

    time.sleep(1)
    return result, title, keys

def get_file(folder_path, key):
    for file in os.listdir(folder_path):
        if file.startswith(key):
            return file


def save_to_xlsx(source_workbook, source_sheet, target_workbook):
    workbook = openpyxl.load_workbook(source_workbook)
    worksheet = workbook[source_sheet]
    i = 1
    print(worksheet.max_column)
    column_end = worksheet.max_column
    while i <= column_end:
        print(worksheet.cell(9, i).value)
        url = worksheet.cell(9, i).value
        if url != None:
            result, title, keys = fetch_amazon_data(url)
            if result:
                print(f"商品标题: {title.text}")
                worksheet.cell(10, i, title.text)
                key_list = []
                for item in keys:
                    if item.text != "":
                        key_list.append(item.text)
                print(f"亮点个数:{len(key_list)}")
                start = 0
                key_end = len(key_list)
                while start < key_end:
                    print(f"第{start}个亮点:{key_list[start]}")
                    worksheet.cell(start + 11, i, key_list[start])
                    start += 1

            else:
                print("程序异常,请查询原因")

        i = i + 1
    workbook.save(target_workbook)

"""
Q1:<strong>We're sorry but store-ui doesn't work properly without JavaScript enabled. Please enable it to continue.</strong>
A1:忽略
"""

def get_img_from_row(source_workbook,source_worksheet,row):
    workbook = openpyxl.load_workbook(source_workbook,data_only=True)
    worksheet = workbook[source_worksheet]
    sku = worksheet[f"BL{row}"].value
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
    chrome_driver = webdriver.Chrome(options=options)
    chrome_driver.get("http://erpx.ksold.ltd:18085/store/product/walmart/template/add")
    # time.sleep(1)
    sku_input = chrome_driver.find_element(By.CSS_SELECTOR,".flex .el-input.el-input--default .el-input__inner")
    sku_input.send_keys(sku)
    sku_label = chrome_driver.find_element(By.CSS_SELECTOR,".w-lg-50p .el-form-item.is-required.el-form-item--default .el-form-item__label")
    sku_label.click()
    time.sleep(2)
    imgs = chrome_driver.find_elements(By.CSS_SELECTOR,"img")
    # time.sleep(20)
    chrome_driver.execute_script("window.open('');")
    for img in imgs:
        src_value = img.get_attribute("src")
        if src_value.startswith("data:image"):
            time.sleep(2)
        else:
            print(src_value)
            chrome_driver.switch_to.window(chrome_driver.window_handles[1])
            chrome_driver.get(src_value)

            img_element = chrome_driver.find_element(By.CSS_SELECTOR,"img")
            action = ActionChains(chrome_driver)
            action.context_click(img_element).perform()
            # time.sleep(3)
            # plugin_option_xpath = "//div[contains(text(),'图片另存为...')]"
            # #在AliPrice上以图搜同款
            # plugin_option_element = chrome_driver.find_element(By.XPATH,plugin_option_xpath)
            # plugin_option_element.click()
            # time.sleep(3)
            # for _ in range(7):
            # action.send_keys(Keys.ARROW_DOWN).perform()
            # time.sleep(3)
            # action.send_keys(Keys.ENTER).perform()

            break

def add_template_from_xlsx(lb,sku,title,brand,store_account,category,code_type):
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
    chrome_driver = webdriver.Chrome(options=options)
    chrome_driver.get("http://erpx.ksold.ltd:18085/store/product/walmart/template/add")

    time.sleep(1)
    weight = chrome_driver.find_element(By.CSS_SELECTOR,".el-form-item__content .el-input.el-input--default.el-input-group.el-input-group--append .el-input__inner")
    weight.send_keys(lb)


    sku_input = chrome_driver.find_element(By.CSS_SELECTOR,".flex .el-input.el-input--default .el-input__inner")
    sku_input.send_keys(sku)
    sku_label = chrome_driver.find_element(By.CSS_SELECTOR,".w-lg-50p .el-form-item.is-required.el-form-item--default .el-form-item__label")
    sku_label.click()
    time.sleep(1)
    #产品标题有问题
    """   """
    inputs = chrome_driver.find_elements(By.CSS_SELECTOR,".el-form-item__content .el-input.el-input--default .el-input__inner")
    # print(len(inputs))
    for input in inputs:
        if input.get_attribute("placeholder") == "请输入产品名称":
            time.sleep(2)
            input.clear()
            time.sleep(1)
            # input.send_keys(title)
            #send_keys()输入字符串,会将字符串里的空格自动去掉
            input.send_keys(title.replace(" ","*"))
            break


    #品牌有问题
    for input in inputs:
        if input.get_attribute("placeholder") == "请输入品牌名称":
            input.clear()
            input.send_keys(brand)
            break



    #下拉列表框最后执行
    """  """
    #el-select w100p el-select--default
    divs = chrome_driver.find_elements(By.CSS_SELECTOR,".el-select.w100p.el-select--default")
    # print(len(divs))
    for div in divs:
        input = div.find_element(By.CSS_SELECTOR,"input")
        print(input.get_attribute("placeholder"))
        div.click()

        time.sleep(1)
        spans = chrome_driver.find_elements(By.CSS_SELECTOR, ".el-select-dropdown.el-popper .el-scrollbar ul li span")
        # print(len(spans))

        for span in spans:
            if len(span.text) != 0:
                print(span.text)
                if span.text == store_account or span.text == category or span.text == code_type:
                    chrome_driver.execute_script("arguments[0].click();", span)



    sub_button = chrome_driver.find_element(By.CSS_SELECTOR,".app-main .submit-button .fixed-button button")
    print(sub_button.text)
    sub_button.click()





if __name__ == "__main__":

    workbook = openpyxl.load_workbook(get_file(os.getcwd(),"Python-草稿"),data_only=True)
    worksheet = workbook["add_template"]
    print(worksheet.max_row)
    # cycle_times = 101 if worksheet.max_row > 101 else worksheet.max_row
    # store_account = "jingqiadg@163.com"
    # category = "Clothing"
    # code_type = "EAN"
    file = open("add_template.txt","r",encoding="UTF-8")
    lines = file.readlines()
    for line in lines:
        if "#" not in line:
            # print(line.split(",")[0])
            # print(line.split(",")[1])
            # print(line.split(",")[2])
            # print(len(line.split(",")[2]))
            self_cycle_time = line.split(",")[0]
            # print("self_cycle_time:"+self_cycle_time)
            lb_column = line.split(",")[1]
            # print("lb_column:"+lb_column)
            sku_column = line.split(",")[2]
            # print("sku_column:"+sku_column)
            title_column = line.split(",")[3]
            # print("title_column:"+title_column)
            brand_column = line.split(",")[4]
            # print("brand_column:"+brand_column)
            store_account = line.split(",")[5]
            # print("store_account:"+store_account)
            category = line.split(",")[6]
            # print("category:"+category)
            code_type = line.split(",")[7].rstrip()
            # print("code_type:"+code_type)
            # print(len(code_type))
    # cycle_times = int(self_cycle_time) if self_cycle_time != "" else worksheet.max_row
    cycle_times = int(self_cycle_time) + 1 if self_cycle_time else worksheet.max_row
    print("cycle_times:"+str(cycle_times))

    for i in range(2,cycle_times+1):
        print(f"循环次数:{i-1}")
        lb = worksheet[f"{lb_column}{i}"].value
        sku = worksheet[f"{sku_column}{i}"].value
        title = worksheet[f"{title_column}{i}"].value
        print(f"产品名称:{title}")
        brand = worksheet[f"{brand_column}{i}"].value
        add_template_from_xlsx(lb,sku,title,brand,store_account,category,code_type)
    """ """

"""
Q1:ChromeDriver only supports characters in the BMP.从网页爬取的标题含有非BMP字符,无法将该标题通过chromedriver录入到谷歌浏览器打开的网页中

"""


"""  
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
    chrome_driver = webdriver.Chrome(options=options)
    chrome_driver.get("http://erpx.ksold.ltd:18085/store/product/walmart/template/add")



    weight = chrome_driver.find_element(By.CSS_SELECTOR,".el-form-item__content .el-input.el-input--default.el-input-group.el-input-group--append .el-input__inner")
    weight.send_keys("1.55")

    # time.sleep(1)
    sku_input = chrome_driver.find_element(By.CSS_SELECTOR,".flex .el-input.el-input--default .el-input__inner")
    sku_input.send_keys("ZNP250213001")
    sku_label = chrome_driver.find_element(By.CSS_SELECTOR,".w-lg-50p .el-form-item.is-required.el-form-item--default .el-form-item__label")
    sku_label.click()
    time.sleep(1)
    #产品标题有问题
    """   """
    inputs = chrome_driver.find_elements(By.CSS_SELECTOR,".el-form-item__content .el-input.el-input--default .el-input__inner")
    print(len(inputs))
    for input in inputs:
        if input.get_attribute("placeholder") == "请输入产品名称":
            input.clear()
            text = 'Mens Valentines Shirt Love Heart Print Shirts Casual Short Sleeve Vacation Shirts Button Down Hawaiian Beach Shirt'
            # for char in text.split(" "):
            #     input.send_keys(f"{char}" "")
            #     time.sleep(1)
            input.send_keys(text.replace(" ","*"))
            break


    #品牌有问题
    for input in inputs:
        if input.get_attribute("placeholder") == "请输入品牌名称":
            input.clear()
            input.send_keys("Momihoom")
            break



    #下拉列表框最后执行
    """  """
    #el-select w100p el-select--default
    divs = chrome_driver.find_elements(By.CSS_SELECTOR,".el-select.w100p.el-select--default")
    #el-select-dropdown el-popper
    # divs = chrome_driver.find_elements(By.CSS_SELECTOR, ".el-select-dropdown.el-popper")
    print(len(divs))
    for div in divs:
        input = div.find_element(By.CSS_SELECTOR,"input")
        print(input.get_attribute("placeholder"))
        div.click()
        # chrome_driver.execute_script("arguments[0].click();", div)
        # time.sleep(2)
        # spans = chrome_driver.find_elements(By.XPATH,"//div[@class='el-select-dropdown__wrap el-scrollbar__wrap']//ul")
        spans = chrome_driver.find_elements(By.CSS_SELECTOR, ".el-select-dropdown.el-popper .el-scrollbar ul li span")
        print(len(spans))
        num = 0
        for span in spans:
            if len(span.text) != 0:
                print(span.text)
                # span.click()
                if span.text =="jingqiadg@163.com" or span.text == "Clothing" or span.text == "EAN":
                    chrome_driver.execute_script("arguments[0].click();", span)


    sub_button = chrome_driver.find_element(By.CSS_SELECTOR,".app-main .submit-button .fixed-button button")
    print(sub_button.text)
    sub_button.click()
"""






