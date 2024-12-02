import openpyxl

xlsx_workbook = openpyxl.load_workbook("Walmart产品1731657481686.xlsx")
#print(xlsx_workbook.sheetnames)
xlsx_sheet = xlsx_workbook["Sheet1"]
#print(xlsx_sheet.max_row)

ItemIdSet = set()
SKUSet = set()

for i in range(2,xlsx_sheet.max_row + 1):
    ItemIdSet.add(xlsx_sheet["A"f"{i}"].value)

for i in ItemIdSet:
    for j in range(2, xlsx_sheet.max_row + 1):
        if i == xlsx_sheet["A"f"{j}"].value:
            print(i,xlsx_sheet["C"f"{j}"].value)
            break






"""  
for j  in ItemIdSet:
    print(j)
"""









