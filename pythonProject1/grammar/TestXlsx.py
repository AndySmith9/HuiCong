import openpyxl
"""
解决问题:
Q1:openpyxl如何判断空单元格?
Q2:openpyxl如何将单元格置空?
"""


xlsx_workbook = openpyxl.load_workbook("1.xlsx")
xlsx_sheet = xlsx_workbook["Sheet1"]
print(xlsx_sheet.max_row)
for i in range(1,98):
    if xlsx_sheet.cell(i+1, 1).value == None:
        print(i+1,xlsx_sheet.cell(i+1, 1).value)

print(xlsx_sheet.cell(4,2).value)
xlsx_sheet.cell(4,2).value = None
print(xlsx_sheet.cell(4,2).value)
xlsx_workbook.save("result.xlsx")