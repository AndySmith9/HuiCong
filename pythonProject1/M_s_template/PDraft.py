import xlrd
import openpyxl
import os



def get_xlsx_rows_equal_xls(source_xls,source_sheet,target_xlsx,target_sheet,save_xlsx,target_start):
    xls_workbook = xlrd.open_workbook(source_xls)
    xls_sheet = xls_workbook.sheet_by_name(source_sheet)
    print(xls_sheet.nrows)

    xlsx_workbook = openpyxl.load_workbook(target_xlsx)
    print(xlsx_workbook.sheetnames)
    xlsx_sheet = xlsx_workbook[target_sheet]
    print(xlsx_sheet.max_row)
    start = xls_sheet.nrows - 2 + 1 + target_start
    xlsx_sheet.delete_rows(start, xlsx_sheet.max_row - start + 1)

    xlsx_workbook.save(save_xlsx)

def copy_xls_to_xlsx(source_xls,source_sheet,source_column,target_xlsx,target_sheet,target_column,target_start):
    xls_workbook = xlrd.open_workbook(source_xls)
    xls_sheet = xls_workbook.sheet_by_name(source_sheet)
    print(xls_sheet.nrows)


    source = []
    for i in range(1, xls_sheet.nrows):
        print(xls_sheet.cell(i, source_column).value)
        source.append(xls_sheet.cell(i, source_column).value)

    xlsx_workbook = openpyxl.load_workbook(target_xlsx)
    print(xlsx_workbook.sheetnames)
    xlsx_sheet = xlsx_workbook[target_sheet]
    print(xlsx_sheet.max_row)

    for index, value in enumerate(source):
        print(index, value)
        row = index + target_start
        print(row)
        print(target_column)
        # xlsx_sheet.cell(row ,1,value)
        # xlsx_sheet['A'+f'{row}'].value = value
        print(f'{target_column}{row}')
        if value == "":
            xlsx_sheet[f'{target_column}{row}'].value = None
        else:
            xlsx_sheet[f'{target_column}{row}'].value = value
    xlsx_workbook.save(target_xlsx)

def get_file_name(folderPath,key):
    for file in os.listdir(folderPath):
        if file.startswith(key):
            return file


# get_xlsx_rows_equal_xls("1.xls","Sheet1","草稿-2WXX20240826.xlsx","Sheet0","Python-草稿-2WXX20240826.xlsx",2)

getEqualRowsDraft = open("getEqualRowsDraft.txt","r",encoding="UTF-8")

for line1 in getEqualRowsDraft.readlines():
    if "#" not in line1:
        print(line1)
        print(line1.split(",")[0])
        print(line1.split(",")[1])
        print(get_file_name(os.getcwd(), line1.split(",")[2]))
        print(line1.split(",")[3])
        print("Python-" + get_file_name(os.getcwd(), line1.split(",")[2]))
        print(int(line1.split(",")[4]))
        get_xlsx_rows_equal_xls(line1.split(",")[0], line1.split(",")[1], get_file_name(os.getcwd(), line1.split(",")[2]), line1.split(",")[3], "Python-" + get_file_name(os.getcwd(), line1.split(",")[2]), int(line1.split(",")[4]))

getEqualRowsDraft.close()


file = open("PDraft.txt", "r", encoding="UTF-8")
for line in file.readlines():
    if "#" not in line:
        print(line)
        print(line.split(",")[0])
        print(line.split(",")[1])
        print(line.split(",")[2])
        print(get_file_name(os.getcwd(), line.split(",")[3]))
        print(line.split(",")[4])
        print(line.split(",")[5])
        #print(line.split(",")[5].replace("\n",""))
        print(line.split(",")[6])
        copy_xls_to_xlsx(line.split(",")[0], line.split(",")[1], int(line.split(",")[2]), get_file_name(os.getcwd(), line.split(",")[3]), line.split(",")[4], line.split(",")[5],int(line.split(",")[6]))

file.close()


"""  
#SKU
copy_xls_to_xlsx("1.xls","Sheet1",0,"Python-草稿-2WXX20240826.xlsx","Sheet0","A")
#父SKU 单属性不用复制
#copy_xls_to_xlsx("1.xls","Sheet1",1,"Python-草稿-2WXX20240826.xlsx","Sheet0","E")
#产品标题
copy_xls_to_xlsx("1.xls","Sheet1",5,"Python-草稿-2WXX20240826.xlsx","Sheet0","BN")
#产品名称
copy_xls_to_xlsx("1.xls","Sheet1",6,"Python-草稿-2WXX20240826.xlsx","Sheet0","BP")
#沃尔玛描述
copy_xls_to_xlsx("1.xls","Sheet1",7,"Python-草稿-2WXX20240826.xlsx","Sheet0","N")
#特性标签
copy_xls_to_xlsx("1.xls","Sheet1",8,"Python-草稿-2WXX20240826.xlsx","Sheet0","AF")
#颜色
copy_xls_to_xlsx("1.xls","Sheet1",9,"Python-草稿-2WXX20240826.xlsx","Sheet0","AG")
#尺寸
copy_xls_to_xlsx("1.xls","Sheet1",10,"Python-草稿-2WXX20240826.xlsx","Sheet0","AH")
#英文材质
copy_xls_to_xlsx("1.xls","Sheet1",12,"Python-草稿-2WXX20240826.xlsx","Sheet0","AJ")
#中文材质
copy_xls_to_xlsx("1.xls","Sheet1",13,"Python-草稿-2WXX20240826.xlsx","Sheet0","AK")
#SKU价(￥)
copy_xls_to_xlsx("1.xls","Sheet1",14,"Python-草稿-2WXX20240826.xlsx","Sheet0","AL")
#重量(g)
copy_xls_to_xlsx("1.xls","Sheet1",18,"Python-草稿-2WXX20240826.xlsx","Sheet0","AM")
#状态
copy_xls_to_xlsx("1.xls","Sheet1",19,"Python-草稿-2WXX20240826.xlsx","Sheet0","AU")
#代理链接1
copy_xls_to_xlsx("1.xls","Sheet1",21,"Python-草稿-2WXX20240826.xlsx","Sheet0","BA")
#代理链接2
copy_xls_to_xlsx("1.xls","Sheet1",22,"Python-草稿-2WXX20240826.xlsx","Sheet0","BB")
#代理链接3
copy_xls_to_xlsx("1.xls","Sheet1",23,"Python-草稿-2WXX20240826.xlsx","Sheet0","BC")
#代理链接4
copy_xls_to_xlsx("1.xls","Sheet1",24,"Python-草稿-2WXX20240826.xlsx","Sheet0","BD")
#代理链接5
copy_xls_to_xlsx("1.xls","Sheet1",25,"Python-草稿-2WXX20240826.xlsx","Sheet0","BE")
#代理链接6
copy_xls_to_xlsx("1.xls","Sheet1",26,"Python-草稿-2WXX20240826.xlsx","Sheet0","BF")
#代理链接7
copy_xls_to_xlsx("1.xls","Sheet1",27,"Python-草稿-2WXX20240826.xlsx","Sheet0","BG")
#代理链接8
copy_xls_to_xlsx("1.xls","Sheet1",28,"Python-草稿-2WXX20240826.xlsx","Sheet0","BH")
#代理链接9
copy_xls_to_xlsx("1.xls","Sheet1",29,"Python-草稿-2WXX20240826.xlsx","Sheet0","BI")
#代理链接100*100缩率图
copy_xls_to_xlsx("1.xls","Sheet1",30,"Python-草稿-2WXX20240826.xlsx","Sheet0","BJ")
"""