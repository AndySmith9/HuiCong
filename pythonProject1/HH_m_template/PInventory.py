import openpyxl
import xlwt


def get_equal_rows_file(source_file,source_sheet_name,target_file,target_sheet_name,targetStart,save_file):
    source_workbook = openpyxl.load_workbook(source_file, data_only=True)
    source_sheet1 = source_workbook[source_sheet_name]
    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet1 = target_workbook[target_sheet_name]
    start = source_sheet1.max_row - 2 + 1 + targetStart
    length = target_sheet1.max_row - start + 1
    target_sheet1.delete_rows(start, length)
    target_workbook.save(save_file)

def copy_xlsx_to_xlsx(source_file,source_sheet,source_columns,target_file,target_sheet,target_columns,targetStart):
    source_workbook = openpyxl.load_workbook(source_file, data_only=True)
    print(source_workbook.sheetnames)
    source_worksheet1 = source_workbook[source_sheet]
    print(source_worksheet1.dimensions)
    source = []
    for i in range(1, source_worksheet1.max_row):
        print(source_worksheet1[f'{source_columns}' + str(i + 1)].value)
        source.append(source_worksheet1[f'{source_columns}' + str(i + 1)].value)

    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet1 = target_workbook[target_sheet]
    for index, value in enumerate(source):
        print(value)
        target_sheet1[f'{target_columns}' + str(index + targetStart)].value = value
    target_workbook.save(target_file)


def copy_xlsx_to_xls(sourceSheet,sourceRow,sourceColumn,targetSheet,targetRow,targetColumn):
    source = []
    for i in range(sourceRow, sourceSheet.max_row + 1):
        print(sourceSheet[f"{sourceColumn}{i}"].value)
        source.append(sourceSheet[f"{sourceColumn}{i}"].value)
    for index, value in enumerate(source):
        targetSheet.write(index + targetRow, targetColumn, value)
def inventory():
    xls_workbook = xlwt.Workbook()
    sheet1 = xls_workbook.add_sheet("Sheet1")
    title = ("SKU", "平台SKU")
    for index, value in enumerate(title):
        sheet1.write(0, index, value)
    xlsx_workbook = openpyxl.load_workbook("Python-草稿-2WXX20240826.xlsx", data_only=True)
    xlsx_sheet = xlsx_workbook["Children"]

    copy_xlsx_to_xls(xlsx_sheet, 2, "A", sheet1, 1, 0)
    copy_xlsx_to_xls(xlsx_sheet, 2, "F", sheet1, 1, 1)

    xls_workbook.save("Python-Import.xls")


#库存
get_equal_rows_file("Python-草稿-2WXX20240826.xlsx","Children","库存-2WXX20240826.xlsx","ExampleSheet",3,"Python-库存-2WXX20240826.xlsx")
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'F',"Python-库存-2WXX20240826.xlsx","ExampleSheet","A",3)

#运费
get_equal_rows_file("Python-草稿-2WXX20240826.xlsx","Children","运费-2WXX20240826.xlsx","Precise Delivery",7,"Python-运费-2WXX20240826.xlsx")
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'F',"Python-运费-2WXX20240826.xlsx","Precise Delivery","D",7)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'AR',"Python-运费-2WXX20240826.xlsx","Precise Delivery","E",7)


#促销
get_equal_rows_file("Python-草稿-2WXX20240826.xlsx","Children","促销-2WXX20240826.xlsx","Price & Promotion",7,"Python-促销-2WXX20240826.xlsx")
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'F',"Python-促销-2WXX20240826.xlsx","Price & Promotion","D",7)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'AO',"Python-促销-2WXX20240826.xlsx","Price & Promotion","F",7)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'AQ',"Python-促销-2WXX20240826.xlsx","Price & Promotion","J",7)

#Lag-Time   行首格式一样,不用修复了
get_equal_rows_file("Python-草稿-2WXX20240826.xlsx","Children","Lag-Time-2WXX20240826.xlsx","Sheet1",3,"Python-Lag-Time-2WXX20240826.xlsx")
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Children",'F',"Python-Lag-Time-2WXX20240826.xlsx","Sheet1","A",3)


#绑定自定义SKU   ERP可以识别,不用修复

inventory()



