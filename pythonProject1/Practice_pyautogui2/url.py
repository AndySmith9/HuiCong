import time
from urllib.error import HTTPError


from selenium import webdriver
from selenium.common import UnexpectedAlertPresentException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import os
import openpyxl
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import sys
import pyautogui


def fetch_amazon_data(url):
    result = False
    title = ""
    keys = []
    try:

        options = Options()
        """
        cmd窗口执行以下命令,需要提前创建文件夹
        chrome.exe --remote-debugging-port=9527 --user-data-dir="F:\selenium\AutomationProfile"
        """
        options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
        driver = webdriver.Chrome(options=options)
        driver.execute_cdp_cmd(
            'Page.addScriptToEvaluateOnNewDocument',
            {'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'}
        )
        driver.get(url)


        """  
        location = driver.find_element(By.CSS_SELECTOR,"#glow-ingress-block")
        location.click()
        GLUXZipUpdateInput = driver.find_element(By.CSS_SELECTOR,"#GLUXZipUpdateInput")
        GLUXZipUpdateInput.send_keys("90091")
        
    
        change = driver.find_element(By.CSS_SELECTOR,"#GLUXChangePostalCodeLink")
        change.click()
        value = driver.find_element(By.CSS_SELECTOR,"#GLUXZipConfirmationValue")
        value.send_keys("90091")
        apply = driver.find_element(By.CSS_SELECTOR, "#GLUXZipUpdate-announce")
        apply.click()
        """
        title = driver.find_element(By.ID, "productTitle")
        keys = driver.find_elements(By.CSS_SELECTOR, "#feature-bullets span")
        result = True
        #driver.close()
    except HTTPError:
        print("HTTPError")
    except Exception:
        print("请检查url:"+url)

    time.sleep(1)
    return result, title, keys

def get_file(folder_path, key):
    for file in os.listdir(folder_path):
        if file.startswith(key):
            return file


def save_to_xlsx(source_workbook, source_sheet, target_workbook):
    workbook = openpyxl.load_workbook(source_workbook)
    worksheet = workbook[source_sheet]
    i = 1
    print(worksheet.max_column)
    column_end = worksheet.max_column
    while i <= column_end:
        print(worksheet.cell(9, i).value)
        url = worksheet.cell(9, i).value
        if url != None:
            result, title, keys = fetch_amazon_data(url)
            if result:
                print(f"商品标题: {title.text}")
                worksheet.cell(10, i, title.text)
                key_list = []
                for item in keys:
                    if item.text != "":
                        key_list.append(item.text)
                print(f"亮点个数:{len(key_list)}")
                start = 0
                key_end = len(key_list)
                while start < key_end:
                    print(f"第{start}个亮点:{key_list[start]}")
                    worksheet.cell(start + 11, i, key_list[start])
                    start += 1

            else:
                print("程序异常,请查询原因")

        i = i + 1
    workbook.save(target_workbook)

"""
Q1:<strong>We're sorry but store-ui doesn't work properly without JavaScript enabled. Please enable it to continue.</strong>
A1:忽略
"""

def get_img_from_row(source_workbook,source_worksheet,row):
    workbook = openpyxl.load_workbook(source_workbook,data_only=True)
    worksheet = workbook[source_worksheet]
    sku = worksheet[f"BL{row}"].value
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
    chrome_driver = webdriver.Chrome(options=options)
    chrome_driver.get("http://erpx.ksold.ltd:18085/store/product/walmart/template/add")
    #time.sleep(1)
    sku_input = chrome_driver.find_element(By.CSS_SELECTOR,".flex .el-input.el-input--default .el-input__inner")
    sku_input.send_keys(sku)
    sku_label = chrome_driver.find_element(By.CSS_SELECTOR,".w-lg-50p .el-form-item.is-required.el-form-item--default .el-form-item__label")
    sku_label.click()
    time.sleep(2)
    imgs = chrome_driver.find_elements(By.CSS_SELECTOR,"img")
    # time.sleep(20)
    chrome_driver.execute_script("window.open('');")
    for img in imgs:
        src_value = img.get_attribute("src")
        if src_value.startswith("data:image"):
            time.sleep(2)
        else:
            print(src_value)
            chrome_driver.switch_to.window(chrome_driver.window_handles[1])
            chrome_driver.get(src_value)

            img_element = chrome_driver.find_element(By.CSS_SELECTOR,"img")
            action = ActionChains(chrome_driver)
            action.context_click(img_element).perform()
            # time.sleep(3)
            # plugin_option_xpath = "//div[contains(text(),'图片另存为...')]"
            # #在AliPrice上以图搜同款
            # plugin_option_element = chrome_driver.find_element(By.XPATH,plugin_option_xpath)
            # plugin_option_element.click()
            # time.sleep(3)
            # for _ in range(7):
            # action.send_keys(Keys.ARROW_DOWN).perform()
            # time.sleep(3)
            # action.send_keys(Keys.ENTER).perform()
            pyautogui.leftClick(990,778)



            break




if __name__ == "__main__":
    if len(sys.argv) > 1:
        for arg in sys.argv[1:]:
            print(f"接收参数为:{arg}")
            get_img_from_row(get_file(os.getcwd(), "Python-草稿"), "Sheet0", arg)
    else:
        print("未接收到参数")


















