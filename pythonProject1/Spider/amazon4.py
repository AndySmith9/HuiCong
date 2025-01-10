import os
import time
from urllib.error import HTTPError
import openpyxl
from pyquery import PyQuery

def fetch_amazon_data(url):
    result = False
    title = ""
    price = []
    try:
        doc = PyQuery(url=url)
        # 提取商品标题
        title = doc('#productTitle').text()
        # 提取商品价格
        price = doc('.a-unordered-list.a-vertical.a-spacing-mini span').items()
        result = True
    except HTTPError:
        print("HTTPError")
    #time.sleep(1)
    return result, title, price

def get_file(folder_path, key):
    for file in os.listdir(folder_path):
        if file.startswith(key):
            return file


def save_to_xlsx(source_workbook, source_sheet, target_workbook):
    workbook = openpyxl.load_workbook(source_workbook)
    worksheet = workbook[source_sheet]
    i = 1
    print(worksheet.max_column)
    end = worksheet.max_column
    while i <= end:
        print(worksheet.cell(9, i).value)
        url = worksheet.cell(9, i).value
        result, title, price = fetch_amazon_data(url)
        if result:
            print(f"商品标题: {title}")
            worksheet.cell(10, i, title)
            key_list = []
            for item in price:
                key_list.append(item.text())
            print(len(key_list))
            start = 0
            end = len(key_list)
            while start < end:
                print(key_list[start])
                worksheet.cell(start + 11, i, key_list[start])
                start += 1

        else:
            print("程序异常,请查询原因")
        i = i + 1
    workbook.save(target_workbook)

if __name__ == "__main__":
    save_to_xlsx(get_file(os.getcwd(), "Python-草稿"), "Key", get_file(os.getcwd(), "Python-草稿"))


"""  
file = get_file(os.getcwd(), "1Python-草稿")
if file == None:
    print(None)
else:
    print(file)
"""





"""   
workbook = openpyxl.load_workbook("Python-草稿-2WXX20250106.xlsx")
worksheet = workbook["Key"]
i = 1
print(worksheet.max_column)
end = worksheet.max_column
while i <= end:
    print(worksheet.cell(9,i).value)
    url = worksheet.cell(9,i).value
    result, title, price = fetch_amazon_data(url)
    if result:
        print(f"商品标题: {title}")
        worksheet.cell(10,i,title)
        #print(f"商品价格: {price}" + ",长度为" + f"{len(list(price))}")
        key_list = []
        for item in price:
            key_list.append(item.text())
        print(len(key_list))
        start = 0
        end = len(key_list)
        while start < end:
            print(key_list[start])
            worksheet.cell(start+11,i,key_list[start])
            start += 1

    else:
        print("程序异常,请查询原因")
    i = i + 1
workbook.save("result.xlsx")
"""


"""     
url = "https://www.amazon.com/dp/B0B2LRRH1K"
result, title, price = fetch_amazon_data(url)
if result:
    print(f"商品标题: {title}")
    print(f"商品价格: {price}"+",长度为"+f"{len(list(price))}")
    for item in price:
        print(item.text())
else:
    print("程序异常,请查询原因")
"""



