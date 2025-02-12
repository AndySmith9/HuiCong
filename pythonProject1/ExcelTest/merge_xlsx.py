import openpyxl


workbook1 = openpyxl.load_workbook("运费-关税-1.xlsx",data_only=True)
worksheet1 = workbook1.active
workbook2 = openpyxl.load_workbook("运费-关税-2.xlsx",data_only=True)
worksheet2 = workbook2.active

for row in worksheet2.iter_rows(min_row=7):
    print(row)
    row_with_values = [cell.value for cell in row]
    print(row_with_values)
    worksheet1.append(row_with_values)

workbook1.save("merged.xlsx")

"""
raise ValueError("Cells cannot be copied from other worksheets")
https://stackoverflow.com/questions/68798047/cells-cannot-be-copied-from-other-worksheets
"""


