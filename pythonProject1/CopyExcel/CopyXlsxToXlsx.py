import openpyxl

"""
source_workbook = openpyxl.load_workbook("草稿-20241105.xlsx",data_only=True)
print(source_workbook.sheetnames)
print(source_workbook.worksheets)
source_sheet1 = source_workbook.worksheets[0]
print(source_sheet1.max_row,source_sheet1.max_column)
source_sheet2 = source_workbook[source_workbook.sheetnames[0]]
print(source_sheet2.max_row,source_sheet2.max_column)


target_workbook = openpyxl.load_workbook("上架表-20241105.xlsx")
print(target_workbook.sheetnames)
print(target_workbook.worksheets)
#target_sheet1 = target_workbook.worksheets[1]
target_sheet1 = target_workbook["Tools & Hardware Other"]
print(target_sheet1.max_row,target_sheet1.max_column)
start = source_sheet1.max_row - 2 + 1 + 7
length =  target_sheet1.max_row - start + 1
print(start,length)
target_sheet1.delete_rows(start,length)
target_workbook.save("Python-"+"上架表-20241105.xlsx")

"""

def get_equal_rows_file(source_file,source_sheet_name,target_file,target_sheet_name):
    source_workbook = openpyxl.load_workbook(source_file, data_only=True)
    source_sheet1 = source_workbook[source_sheet_name]
    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet1 = target_workbook[target_sheet_name]
    start = source_sheet1.max_row - 2 + 1 + 7
    length = target_sheet1.max_row - start + 1
    target_sheet1.delete_rows(start, length)
    target_workbook.save("Python-" + target_file)


#get_equal_rows_file("草稿-20241105.xlsx","Sheet0","草稿-20241105.xlsx","Sheet0")
#get_equal_rows_file("草稿-20241105.xlsx","Sheet0","上架表-20241105.xlsx","Tools & Hardware Other")

"""
source_workbook = openpyxl.load_workbook("Python-草稿-20241105.xlsx", data_only=True)
print(source_workbook.sheetnames)
source_worksheet1 = source_workbook["Sheet0"]
print(source_worksheet1.dimensions)
source = []


# print(source_worksheet1["D2"].value)
# print(source_worksheet1.cell(2, 4).value)
# print(source_worksheet1["E2"])
# print(source_worksheet1.cell(2,5))
# print(source_worksheet1["E2"].value)



for i in range(1,source_worksheet1.max_row):
    print(source_worksheet1['E' + str(i+1)].value)
    source.append(source_worksheet1['E' + str(i+1)].value)

target_workbook = openpyxl.load_workbook("Python-上架表-20241105.xlsx")
target_sheet1 = target_workbook["Tools & Hardware Other"]
for index, value in enumerate(source):
    print(value)
    target_sheet1["D"+str(index+7)].value = value
target_workbook.save("Python-上架表-20241105.xlsx")
"""



def copy_xlsx_to_xlsx(source_file,source_sheet,source_columns,target_file,target_sheet,target_columns):
    source_workbook = openpyxl.load_workbook(source_file, data_only=True)
    print(source_workbook.sheetnames)
    source_worksheet1 = source_workbook[source_sheet]
    print(source_worksheet1.dimensions)
    source = []
    for i in range(1, source_worksheet1.max_row):
        print(source_worksheet1[f'{source_columns}' + str(i + 1)].value)
        source.append(source_worksheet1[f'{source_columns}' + str(i + 1)].value)

    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet1 = target_workbook[target_sheet]
    for index, value in enumerate(source):
        print(value)
        target_sheet1[f'{target_columns}' + str(index + 7)].value = value
    target_workbook.save(target_file)



copy_xlsx_to_xlsx("Python-草稿-20241105.xlsx","Sheet0",'E',"Python-上架表-20241105.xlsx","Tools & Hardware Other","D")



