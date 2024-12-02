import xlrd
import openpyxl

"""
xls_workbook = xlrd.open_workbook("1.xls")
xls_sheet = xls_workbook.sheet_by_name(xls_workbook.sheet_names()[0])
print(xls_sheet.nrows)
#print(xls_sheet.cell(1,0))


source = []
for i in range(1,xls_sheet.nrows):
    #    print(i)
    print(xls_sheet.cell(i,0).value)
    source.append(xls_sheet.cell(i,0).value)


xlsx_workbook = openpyxl.load_workbook("草稿-20241105.xlsx")
print(xlsx_workbook.sheetnames)
xlsx_sheet = xlsx_workbook[xlsx_workbook.sheetnames[0]]
print(xlsx_sheet.max_row)
start = xls_sheet.nrows - 2 + 1 + 2
xlsx_sheet.delete_rows(start,xlsx_sheet.max_row - start + 1)
for index,value in enumerate(source):
    print(index, value)
    row = index + 2
    print(row)
    # xlsx_sheet.cell(row ,1,value)
    #xlsx_sheet['A'+f'{row}'].value = value
    xlsx_sheet[f'A{row}'].value = value
xlsx_workbook.save("delete.xlsx")
"""

"""
raise ValueError("Cannot convert {0!r} to Excel".format(value))
ValueError: Cannot convert text:'WHL241028003' to Excel

ValueError: A{row} is not a valid coordinate or range
"""


def get_rows_equal_xlsx(xls_file,xlsx_file):
    xls_workbook = xlrd.open_workbook(xls_file)
    xls_sheet = xls_workbook.sheet_by_name(xls_workbook.sheet_names()[0])
    print(xls_sheet.nrows)

    xlsx_workbook = openpyxl.load_workbook(xlsx_file)
    print(xlsx_workbook.sheetnames)
    xlsx_sheet = xlsx_workbook[xlsx_workbook.sheetnames[0]]
    print(xlsx_sheet.max_row)
    start = xls_sheet.nrows - 2 + 1 + 2
    xlsx_sheet.delete_rows(start, xlsx_sheet.max_row - start + 1)

    xlsx_workbook.save("Python-"+xlsx_file)

def copy_xls_to_xlsx(xls_file,source_columns,xlsx_file,target_columns):
    xls_workbook = xlrd.open_workbook(xls_file)
    xls_sheet = xls_workbook.sheet_by_name(xls_workbook.sheet_names()[0])
    print(xls_sheet.nrows)
    # print(xls_sheet.cell(1,0))

    source = []
    for i in range(1, xls_sheet.nrows):
        #    print(i)
        print(xls_sheet.cell(i, source_columns).value)
        source.append(xls_sheet.cell(i, source_columns).value)

    xlsx_workbook = openpyxl.load_workbook("Python-"+xlsx_file)
    print(xlsx_workbook.sheetnames)
    xlsx_sheet = xlsx_workbook[xlsx_workbook.sheetnames[0]]
    print(xlsx_sheet.max_row)

    for index, value in enumerate(source):
        print(index, value)
        row = index + 2
        print(row)
        print(target_columns)
        # xlsx_sheet.cell(row ,1,value)
        # xlsx_sheet['A'+f'{row}'].value = value
        xlsx_sheet[f'{target_columns}{row}'].value = value
    xlsx_workbook.save("Python-"+xlsx_file)




get_rows_equal_xlsx("1.xls","草稿-20241105.xlsx")
#SKU
copy_xls_to_xlsx("1.xls",0,"草稿-20241105.xlsx","A")
#产品标题
copy_xls_to_xlsx("1.xls",5,"草稿-20241105.xlsx","BN")
#产品名称
copy_xls_to_xlsx("1.xls",6,"草稿-20241105.xlsx","BP")
#沃尔玛描述
copy_xls_to_xlsx("1.xls",7,"草稿-20241105.xlsx","L")
#特性标签
copy_xls_to_xlsx("1.xls",8,"草稿-20241105.xlsx","AD")
#颜色
copy_xls_to_xlsx("1.xls",9,"草稿-20241105.xlsx","AE")
#尺寸
copy_xls_to_xlsx("1.xls",10,"草稿-20241105.xlsx","AF")
#英文材质
copy_xls_to_xlsx("1.xls",12,"草稿-20241105.xlsx","AH")
#中文材质
copy_xls_to_xlsx("1.xls",13,"草稿-20241105.xlsx","AI")
#SKU价(￥)
copy_xls_to_xlsx("1.xls",14,"草稿-20241105.xlsx","AJ")
#重量(g)
copy_xls_to_xlsx("1.xls",18,"草稿-20241105.xlsx","AK")
#状态
copy_xls_to_xlsx("1.xls",19,"草稿-20241105.xlsx","AS")
#代理链接1
copy_xls_to_xlsx("1.xls",21,"草稿-20241105.xlsx","BA")
#代理链接2
copy_xls_to_xlsx("1.xls",22,"草稿-20241105.xlsx","BB")
#代理链接3
copy_xls_to_xlsx("1.xls",23,"草稿-20241105.xlsx","BC")
#代理链接4
copy_xls_to_xlsx("1.xls",24,"草稿-20241105.xlsx","BD")
#代理链接5
copy_xls_to_xlsx("1.xls",25,"草稿-20241105.xlsx","BE")
#代理链接6
copy_xls_to_xlsx("1.xls",26,"草稿-20241105.xlsx","BF")
#代理链接7
copy_xls_to_xlsx("1.xls",27,"草稿-20241105.xlsx","BG")
#代理链接8
copy_xls_to_xlsx("1.xls",28,"草稿-20241105.xlsx","BH")
#代理链接9
copy_xls_to_xlsx("1.xls",29,"草稿-20241105.xlsx","BI")
#代理链接100*100缩率图
copy_xls_to_xlsx("1.xls",30,"草稿-20241105.xlsx","BJ")
"""
"""
#
#
#
#