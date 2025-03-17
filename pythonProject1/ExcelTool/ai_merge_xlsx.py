import openpyxl
from openpyxl.utils import get_column_letter

def copy_cell(source_cell, target_cell):
    """
    复制单元格的值和样式
    """
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = source_cell.font
        target_cell.border = source_cell.border
        target_cell.fill = source_cell.fill
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection
        target_cell.alignment = source_cell.alignment

def merge_excel_files(file_paths, output_file):
    """
    合并多个 Excel 文件
    """
    # 创建一个新的工作簿作为目标文件
    merged_wb = openpyxl.Workbook()
    # 删除默认的工作表
    del merged_wb['Sheet']

    for file_path in file_paths:
        # 打开源文件
        source_wb = openpyxl.load_workbook(file_path)
        for sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[sheet_name]
            # 检查目标工作簿中是否已存在同名工作表
            if sheet_name not in merged_wb.sheetnames:
                # 如果不存在，则创建一个新的工作表
                target_sheet = merged_wb.create_sheet(sheet_name)
            else:
                # 如果存在，则获取该工作表
                target_sheet = merged_wb[sheet_name]

            # 复制列宽
            for col in source_sheet.columns:
                col_letter = get_column_letter(col[0].column)
                target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

            # 获取目标工作表的最大行数，以便追加数据
            max_row = target_sheet.max_row
            for row in source_sheet.iter_rows():
                max_row += 1
                for cell in row:
                    target_cell = target_sheet.cell(row=max_row, column=cell.column)
                    copy_cell(cell, target_cell)

        # 关闭源文件
        source_wb.close()

    # 保存合并后的工作簿
    merged_wb.save(output_file)
    print(f"合并完成，结果保存到 {output_file}")

# 定义要合并的文件路径列表
file_paths = ['file1.xlsx', 'file2.xlsx']
# 定义输出文件路径
output_file = 'merged.xlsx'

# 调用合并函数
merge_excel_files(file_paths, output_file)