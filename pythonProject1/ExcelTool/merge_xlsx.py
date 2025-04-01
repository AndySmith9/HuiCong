import os

import openpyxl

merge_file = open("merge.txt","r",encoding="ANSI")
lines = merge_file.readlines()

"""  
for index,row in enumerate(lines):
    print("index:"+str(index),"row:"+row)
    locals()["workbook"+str(index)] = openpyxl.load_workbook(row.split(",")[0],data_only=True)
    print(locals()["workbook"+str(index)].sheetnames)
    locals()["worksheet"+str(index)] = locals()["workbook"+str(index)].active
    print(locals()["worksheet"+str(index)].title)
"""

"""
配置文件格式:
D:\店铺产品-jingqiadg@163.com\202502\促销-merged.xlsx,
D:\店铺产品-jingqiadg@163.com\202502\20250210\促销-WXX20250210.xlsx,
D:\店铺产品-jingqiadg@163.com\202502\20250211\促销-WXX20250211.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250212\促销-WXX20250212.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250213\促销-WXX20250213.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250214\促销-WXX20250214.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250215\促销-WXX20250215.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250217\促销-WXX20250217.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250219\促销-WXX20250219.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250220\促销-WXX20250220.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250221\促销-WXX20250221.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250224\促销-WXX20250224.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250227\促销-WXX20250227.xlsx,7
D:\店铺产品-jingqiadg@163.com\202502\20250228\促销-WXX20250228.xlsx,7
"""

for i in range(1,len(lines)):
    locals()["workbook"+str(i)] = openpyxl.load_workbook(lines[i].split(",")[0],data_only=True)
    print(locals()["workbook"+str(i)].sheetnames)
    locals()["worksheet"+str(i)] = locals()["workbook"+str(i)].active
    print(locals()["worksheet"+str(i)].title)


for i in range(2,len(lines)):
    print(i)
    for row in locals()["worksheet"+str(i)].iter_rows(min_row=int(lines[i].split(",")[1])):
        print(row)
        row_values = [cell.value for cell in row]
        locals()["worksheet1"].append(row_values)


locals()["workbook1"].save(lines[0].split(",")[0])




merge_file.close()


