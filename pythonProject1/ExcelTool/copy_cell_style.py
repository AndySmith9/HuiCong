import openpyxl





def copy_cell(source_cell, target_cell):
  
    #复制单元格的值和样式
  
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = source_cell.font
        target_cell.border = source_cell.border
        target_cell.fill = source_cell.fill
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection



if __name__ == "__main__":
    print("输出")
    workbook1 = openpyxl.load_workbook("file1.xlsx")
    worksheet1 = workbook1["Sheet0"]
    print("输出")
    print("最大列:"+str(worksheet1.max_column)+",最大行:"+str(worksheet1.max_row))
    """   
    copy_cell(worksheet1["A2"],worksheet1["A13"])
    workbook1.save("result.xlsx")
    """
    if worksheet1["A2"].has_style:
        print("有格式")
        print(worksheet1["A2"].font)
        print(worksheet1["A2"].border)
        print(worksheet1["A2"].fill)
        print(worksheet1["A2"].protection)
        print(worksheet1["A2"].number_format)
        print(worksheet1["A2"].alignment)

    worksheet1["A14"].value = worksheet1["A2"].value
    worksheet1["A14"].border = worksheet1["A2"].border
    """  
    worksheet1["A14"].font = worksheet1["A2"].font
    worksheet1["A14"].fill = worksheet1["A2"].fill
    worksheet1["A14"].protection = worksheet1["A2"].protection
    worksheet1["A14"].number_format = worksheet1["A2"].number_format
    worksheet1["A14"].alignment = worksheet1["A2"].alignment
    """
    workbook1.save("result.xlsx")



