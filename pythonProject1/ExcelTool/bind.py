import openpyxl


def copy_xlsx_to_xlsx_column(source_workbook_name,source_worksheet_name,source_column,source_row,target_workbook_name,target_worksheet_name,target_column,target_row):
    source_workbook = openpyxl.load_workbook(source_workbook_name,data_only=True)
    source_sheet = source_workbook[source_worksheet_name]
    end = source_sheet.max_row + source_row - 1
    print(end)
    source_list = []
    for i in range(source_row,end):
        print(source_sheet[f"{source_column}{i}"].value)
        source_list.append(source_sheet[f"{source_column}{i}"].value)
    print(len(source_list))
    target_workbook = openpyxl.load_workbook(target_workbook_name)
    target_worksheet = target_workbook[target_worksheet_name]
    for index,value in enumerate(source_list):
        target_worksheet[f"{target_column}{index*2+target_row}"].value = value
    target_workbook.save(target_workbook_name)



def xlsx_write_value(source_workbook_name,source_worksheet_name,target_workbook_name,target_worksheet_name,target_column,target_row):
    source_workbook = openpyxl.load_workbook(source_workbook_name,data_only=True)
    source_worksheet = source_workbook[source_worksheet_name]
    rows = source_worksheet.max_row
    end = rows * 2 + 1

    target_workbook = openpyxl.load_workbook(target_workbook_name)
    target_worksheet = target_workbook[target_worksheet_name]
    for i in range(target_row,end,2):
        target_worksheet[f"{target_column}{i}"].value = 4
    target_workbook.save(target_workbook_name)



"""
配置文件格式:
copy_xlsx_to_xlsx_column,E:\HuiCong\pythonProject1\ExcelTest\Python-草稿-2WXX20250207.xlsx,Sheet0,F,2,E:\HuiCong\pythonProject1\ExcelTest\product_bind_import_20250207.xlsx,Sheet1,A,3
copy_xlsx_to_xlsx_column,E:\HuiCong\pythonProject1\ExcelTest\Python-草稿-2WXX20250207.xlsx,Sheet0,L,2,E:\HuiCong\pythonProject1\ExcelTest\product_bind_import_20250207.xlsx,Sheet1,B,3
copy_xlsx_to_xlsx_column,E:\HuiCong\pythonProject1\ExcelTest\Python-草稿-2WXX20250207.xlsx,Sheet0,A,2,E:\HuiCong\pythonProject1\ExcelTest\product_bind_import_20250207.xlsx,Sheet1,B,4
xlsx_write_value,E:\HuiCong\pythonProject1\ExcelTest\Python-草稿-2WXX20250207.xlsx,Sheet0,E:\HuiCong\pythonProject1\ExcelTest\product_bind_import_20250207.xlsx,Sheet1,C,4
注意:Python-草稿-2WXX20250207.xlsx和product_bind_import_20250207.xlsx需要存在,product_bind_import_20250207.xlsx可以新建一个空的xlsx文件,然后改名
"""



readFile = open("bind.txt","r",encoding="ANSI")
for i in readFile.readlines():
    if "copy_xlsx_to_xlsx_column" in i:
        copy_xlsx_to_xlsx_column(i.split(",")[1],i.split(",")[2],i.split(",")[3],int(i.split(",")[4]),i.split(",")[5],i.split(",")[6],i.split(",")[7],int(i.split(",")[8]))
    elif "xlsx_write_value" in i:
        xlsx_write_value(i.split(",")[1],i.split(",")[2],i.split(",")[3],i.split(",")[4],i.split(",")[5],int(i.split(",")[6]))

readFile.close()


