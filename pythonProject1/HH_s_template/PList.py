import openpyxl



def get_xlsx_equal_rows_xlsx(source_xlsx,source_sheet,target_xlsx,target_sheet,save_xlsx,target_start):
    source_workbook = openpyxl.load_workbook(source_xlsx, data_only=True)
    source_sheet1 = source_workbook[source_sheet]
    target_workbook = openpyxl.load_workbook(target_xlsx)
    target_sheet1 = target_workbook[target_sheet]
    start = source_sheet1.max_row - 2 + 1 + target_start
    length = target_sheet1.max_row - start + 1
    target_sheet1.delete_rows(start, length)
    target_workbook.save(save_xlsx)
def copy_xlsx_to_xlsx(source_file,source_sheet,source_column,target_file,target_sheet,target_column):
    source_workbook = openpyxl.load_workbook(source_file, data_only=True)
    print(source_workbook.sheetnames)
    source_worksheet1 = source_workbook[source_sheet]
    print(source_worksheet1.dimensions)
    source = []
    for i in range(1, source_worksheet1.max_row):
        print(source_worksheet1[f'{source_column}' + str(i + 1)].value)
        source.append(source_worksheet1[f'{source_column}' + str(i + 1)].value)

    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet1 = target_workbook[target_sheet]
    for index, value in enumerate(source):
        print(value)
        target_sheet1[f'{target_column}' + str(index + 7)].value = value
    target_workbook.save(target_file)


get_xlsx_equal_rows_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0","上架表-2WXX20240826.xlsx","Clothing","Python-上架表-2WXX20240826.xlsx",7)
"""
openpyxl提供的delete_rows函数,只能清除行内容,不能删除行.被清除行内容的行,还是会被max_row统计到 ?
草稿表xlsx是xls转化来的
"""


#自定义SKU - SKU
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'F',"Python-上架表-2WXX20240826.xlsx","Clothing","D")
#上架标题 - Product Name
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'L',"Python-上架表-2WXX20240826.xlsx","Clothing","E")
#建议价（规则）- Selling Price
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AO',"Python-上架表-2WXX20240826.xlsx","Clothing","H")
#重量（LB）- Shipping Weight (lbs)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AN',"Python-上架表-2WXX20240826.xlsx","Clothing","J")
#代理链接 1(主图) - Main Image URL
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BA',"Python-上架表-2WXX20240826.xlsx","Clothing","K")
#沃尔玛描述 - Site Description
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'N',"Python-上架表-2WXX20240826.xlsx","Clothing","L")
#建议价（规则）- MSRP
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AO',"Python-上架表-2WXX20240826.xlsx","Clothing","BC")
#代理链接 2 - Additional Image URL (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BB',"Python-上架表-2WXX20240826.xlsx","Clothing","BF")
#代理链接 3 - Additional Image URL 1 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BC',"Python-上架表-2WXX20240826.xlsx","Clothing","BG")
#代理链接 4 - Additional Image URL 2 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BD',"Python-上架表-2WXX20240826.xlsx","Clothing","BH")
#代理链接 5 - Additional Image URL 3 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BE',"Python-上架表-2WXX20240826.xlsx","Clothing","BI")
#代理链接 6 - Additional Image URL 4 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BF',"Python-上架表-2WXX20240826.xlsx","Clothing","BJ")
#代理链接 7 - Additional Image URL 5 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BG',"Python-上架表-2WXX20240826.xlsx","Clothing","BK")
#代理链接 8 - Additional Image URL 6 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BH',"Python-上架表-2WXX20240826.xlsx","Clothing","BL")
#代理链接 9 - Additional Image URL 7 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BI',"Python-上架表-2WXX20240826.xlsx","Clothing","BM")
#颜色 - Color (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AG',"Python-上架表-2WXX20240826.xlsx","Clothing","DC")
#亮点1 - Key Features (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'Y',"Python-上架表-2WXX20240826.xlsx","Clothing","DE")
#亮点2 - Key Features 1 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'Z',"Python-上架表-2WXX20240826.xlsx","Clothing","DF")
#亮点3 - Key Features 2 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AA',"Python-上架表-2WXX20240826.xlsx","Clothing","DG")
#亮点4 - Key Features 3 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AB',"Python-上架表-2WXX20240826.xlsx","Clothing","DH")
#亮点5 - Key Features 4 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AC',"Python-上架表-2WXX20240826.xlsx","Clothing","DI")
#亮点6 - Key Features 5 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AD',"Python-上架表-2WXX20240826.xlsx","Clothing","DJ")
#亮点7 - Key Features 6 (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AE',"Python-上架表-2WXX20240826.xlsx","Clothing","DK")
#尺寸 - Clothing Size
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AH',"Python-上架表-2WXX20240826.xlsx","Clothing","DS")
#英文材质 - Material (+)
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'AJ',"Python-上架表-2WXX20240826.xlsx","Clothing","EH")
#自定义SKU - Variant Group ID
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'F',"Python-上架表-2WXX20240826.xlsx","Clothing","EN")
#代理链接100*100缩率图(Formula) - Swatch Image URL
copy_xlsx_to_xlsx("Python-草稿-2WXX20240826.xlsx","Sheet0",'BK',"Python-上架表-2WXX20240826.xlsx","Clothing","ER")





