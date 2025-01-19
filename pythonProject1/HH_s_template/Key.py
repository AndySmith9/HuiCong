import openpyxl
import os

def replace_key(sourceFile,sourceSheet,targetFile,targetSheet):
    source_workbook = openpyxl.load_workbook(sourceFile, data_only=True)
    source_sheet = source_workbook[sourceSheet]
    print(source_sheet.max_column)

    target_workbook = openpyxl.load_workbook(targetFile)
    target_sheet = target_workbook[targetSheet]
    print(target_sheet.max_row)

    for i in range(1, source_sheet.max_column + 1):
        print(source_sheet.cell(1, i).value, source_sheet.cell(8, i).value)
        start = source_sheet.cell(1, i).value
        while start <= source_sheet.cell(8, i).value:
            """ 
            if source_sheet.cell(2, i).value != None:
                target_sheet["BN"f"{start}"].value = source_sheet.cell(2, i).value
            if source_sheet.cell(3, i).value != None:
                target_sheet["AA"f"{start}"].value = source_sheet.cell(3, i).value
            if source_sheet.cell(4, i).value != None:
                target_sheet["AB"f"{start}"].value = source_sheet.cell(4, i).value
            if source_sheet.cell(5, i).value != None:
                target_sheet["AC"f"{start}"].value = source_sheet.cell(5, i).value
            if source_sheet.cell(6, i).value != None:
                target_sheet["AD"f"{start}"].value = source_sheet.cell(6, i).value
            if source_sheet.cell(7, i).value != None:
                target_sheet["AE"f"{start}"].value = source_sheet.cell(7, i).value
            """
            target_sheet["BN"f"{start}"].value = source_sheet.cell(2, i).value
            target_sheet["AA"f"{start}"].value = source_sheet.cell(3, i).value
            target_sheet["AB"f"{start}"].value = source_sheet.cell(4, i).value
            target_sheet["AC"f"{start}"].value = source_sheet.cell(5, i).value
            target_sheet["AD"f"{start}"].value = source_sheet.cell(6, i).value
            target_sheet["AE"f"{start}"].value = source_sheet.cell(7, i).value


            start = start + 1

    target_workbook.save(targetFile)

def get_file_name(folderPath,key):
    for file in os.listdir(folderPath):
        if file.startswith(key):
            return file

#replace_key("Python-草稿-2WXX20240826.xlsx","Key","Python-草稿-2WXX20240826.xlsx","Sheet0")
replace_key(get_file_name(os.getcwd(), "Python-草稿"),"Key",get_file_name(os.getcwd(), "Python-草稿"),"Sheet0")
